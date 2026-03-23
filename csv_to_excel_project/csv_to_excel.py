import pandas as pd

print("Loading CSV file...")

data = pd.read_csv("sales_data.csv")

print("Calculating profit...")

data["Profit"] = data["Sales"] - data["Cost"]

print("Saving Excel file with formatting...")

# Save to Excel with formatting
with pd.ExcelWriter("sales_report.xlsx", engine="openpyxl") as writer:
    data.to_excel(writer, index=False, sheet_name="Report")
    
    workbook = writer.book
    worksheet = writer.sheets["Report"]

    from openpyxl.styles import PatternFill

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # Find Profit column index
    profit_col_index = list(data.columns).index("Profit") + 1

    # Apply red color to negative profit cells
    for row in range(2, len(data) + 2):
        cell = worksheet.cell(row=row, column=profit_col_index)
        if cell.value < 0:
            cell.fill = red_fill

print("Conversion complete with highlighted losses!")