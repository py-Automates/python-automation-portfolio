import pandas as pd

print("Loading Excel dataset...")

data = pd.read_excel("sales_data_large.xlsx")

print("Calculating profit for each row...")

data["Profit"] = data["Sales"] - data["Cost"]

print("Calculating totals...")

total_sales = data["Sales"].sum()
total_cost = data["Cost"].sum()
total_profit = data["Profit"].sum()

print("Total Sales:", total_sales)
print("Total Cost:", total_cost)
print("Total Profit:", total_profit)

print("Saving automated report with formatting...")

with pd.ExcelWriter("sales_profit_report.xlsx", engine="openpyxl") as writer:
    
    data.to_excel(writer, index=False, sheet_name="Report")

    worksheet = writer.sheets["Report"]

    from openpyxl.styles import PatternFill

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    profit_col_index = list(data.columns).index("Profit") + 1

    for row in range(2, len(data) + 2):
        cell = worksheet.cell(row=row, column=profit_col_index)
        if cell.value < 0:
            cell.fill = red_fill
        else:
            cell.fill = green_fill

print("Report generated successfully with highlights!")