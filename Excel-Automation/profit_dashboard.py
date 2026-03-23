import pandas as pd

print("Loading sales data...")

data = pd.read_excel("sales_data_large.xlsx")

print("Creating profit column...")

data["Profit"] = data["Sales"] - data["Cost"]

print("Calculating totals...")

total_sales = data["Sales"].sum()
total_cost = data["Cost"].sum()
total_profit = data["Profit"].sum()

print("Total Sales:", total_sales)
print("Total Cost:", total_cost)
print("Total Profit:", total_profit)

print("Creating product summary...")

summary = data.groupby("Product")[["Sales","Cost","Profit"]].sum()

print("Saving dashboard report with formatting...")

with pd.ExcelWriter("sales_dashboard_report.xlsx", engine="openpyxl") as writer:
    
    data.to_excel(writer, sheet_name="Raw Data", index=False)
    summary.to_excel(writer, sheet_name="Product Summary")

    workbook = writer.book
    ws1 = writer.sheets["Raw Data"]
    ws2 = writer.sheets["Product Summary"]

    from openpyxl.styles import PatternFill

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    # Highlight Raw Data Profit column
    profit_col_index = list(data.columns).index("Profit") + 1

    for row in range(2, len(data) + 2):
        cell = ws1.cell(row=row, column=profit_col_index)
        if cell.value < 0:
            cell.fill = red_fill
        else:
            cell.fill = green_fill

    # Highlight Summary Profit column
    summary_profit_col = list(summary.columns).index("Profit") + 2  # +2 because index column exists

    for row in range(2, len(summary) + 2):
        cell = ws2.cell(row=row, column=summary_profit_col)
        if cell.value < 0:
            cell.fill = red_fill
        else:
            cell.fill = green_fill

print("Dashboard report generated successfully with highlights!")